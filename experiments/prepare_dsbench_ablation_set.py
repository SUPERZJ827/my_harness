import argparse
import csv
import json
import re
import shutil
from pathlib import Path
from zipfile import ZipFile


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DATA_DIR = PROJECT_ROOT / "data"
DEFAULT_TASKSET_FILE = PROJECT_ROOT / "experiments" / "dsbench_task_sets.json"
DEFAULT_DSBENCH_ROOT = Path("/home/zhoujun/DSBench")

PROMPT_TEMPLATE = """Answer the following DSBench data-analysis task using only the files staged in the current directory.

Challenge ID: {challenge_id}
Challenge Name: {challenge_name}
Year: {year}
Question ID: {question_name}

Challenge Background:
{introduction}

Current Question:
{question}

Available support files in the current directory:
{support_files}

Required output:
1. Write the final answer only to a UTF-8 text file named `final_answer.txt`.
2. The content of `final_answer.txt` must contain only the final answer string and nothing else.
3. Use the workbook / image / csv files staged in `./` as needed. Read inputs from `./` and write outputs to `./`.
4. If the question is multiple choice, write only the final choice exactly as the answer format requires.
5. Do not write explanations into `final_answer.txt`.

DSBench workbook-solving requirements:
1. Treat each task as a spreadsheet/data-analysis problem. Do not answer from memory or from the question text alone.
2. Before computing the final answer, inspect the staged workbook files with `openpyxl` and/or `pandas`: list sheet names, sheet dimensions, visible data ranges, and relevant formulas or cell values.
3. If an image file such as a flow chart is staged, use it as task context when the question or introduction references it. Do not ignore image files that are listed as support files.
4. For multiple-choice questions, read the option list from the question text, compute the underlying value or condition, then map the result to exactly one option letter.
5. For free-field questions, write the requested value or name in the format requested by the question; do not write a placeholder, debug string, or explanation.
6. Do not write `Not Applicable` merely because workbook parsing is difficult. Use `Not Applicable` only after confirming that the required workbook/question/support files or requested entity are genuinely absent.
7. If a workbook contains formulas, inspect both formula cells and cached/displayed values where available. Do not assume `pandas.read_excel` alone captures the full workbook logic.
{extra_instructions}
"""


CURATED_PRESETS = {
    "ablation_30": {
        "description": (
            "30-question DSBench data-analysis ablation set spanning three workbook-heavy challenges. "
            "Chosen to balance mixed support files, later-step numeric reasoning, and a mix of MCQ and open answers."
        ),
        "selections": [
            {
                "challenge_id": "00000001",
                "reason": "Long quarterly project-modeling challenge; questions 6-15 stress repeated structured spreadsheet reasoning.",
                "question_names": [
                    "question6",
                    "question7",
                    "question8",
                    "question9",
                    "question10",
                    "question11",
                    "question12",
                    "question13",
                    "question14",
                    "question15",
                ],
            },
            {
                "challenge_id": "00000004",
                "reason": "Mixed workbook + image + csv support; questions 41-50 require combining flow logic with workbook numbers.",
                "question_names": [
                    "question41",
                    "question42",
                    "question43",
                    "question44",
                    "question45",
                    "question46",
                    "question47",
                    "question48",
                    "question49",
                    "question50",
                ],
            },
            {
                "challenge_id": "00000010",
                "reason": "Workbook task with mixed MCQ, free-text, and integer answers; useful for testing whether ablations handle format shifts.",
                "question_names": [
                    "question1",
                    "question2",
                    "question3",
                    "question4",
                    "question5",
                    "question6",
                    "question7",
                    "question8",
                    "question9",
                    "question10",
                ],
            },
        ],
    }
}


def parse_args():
    parser = argparse.ArgumentParser(description="Prepare a larger DSBench ablation task set for DataSciBench")
    parser.add_argument("--preset", default="ablation_30", choices=sorted(CURATED_PRESETS))
    parser.add_argument("--dsbench-root", default=str(DEFAULT_DSBENCH_ROOT))
    parser.add_argument("--source-zip", default=None)
    parser.add_argument("--meta-file", default=None)
    parser.add_argument("--output-data-dir", default=str(DEFAULT_OUTPUT_DATA_DIR))
    parser.add_argument("--taskset-file", default=str(DEFAULT_TASKSET_FILE))
    parser.add_argument("--task-prefix", default="dsbench_da")
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def answer_to_text(answer):
    if isinstance(answer, (dict, list)):
        return json.dumps(answer, ensure_ascii=False, sort_keys=True)
    return str(answer)


def is_multiple_choice_answer(answer) -> bool:
    text = answer_to_text(answer).strip()
    return len(text) == 1 and text.isalpha()


def sanitize_name(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in value)


def safe_cell_value(value, max_len: int = 120):
    text = "" if value is None else str(value)
    text = " ".join(text.split())
    return text[:max_len]


def generate_workbook_summary(task_dir: Path) -> str | None:
    workbook_paths = sorted(list(task_dir.glob("*.xlsx")) + list(task_dir.glob("*.xlsm")) + list(task_dir.glob("*.xls")))
    if not workbook_paths:
        return None
    lines = [
        "# Workbook Summary",
        "",
        "This file was generated by the DSBench harness to expose workbook structure before agent execution.",
        "Use it as a navigation aid; inspect the original workbook for exact calculations when needed.",
        "",
    ]
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        lines.append(f"Could not import openpyxl: {exc}")
        output = task_dir / "workbook_summary.md"
        output.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return output.name

    for workbook_path in workbook_paths:
        lines.extend([f"## {workbook_path.name}", ""])
        try:
            wb_values = load_workbook(workbook_path, data_only=True, read_only=True)
            wb_formulas = load_workbook(workbook_path, data_only=False, read_only=True)
        except Exception as exc:
            lines.extend([f"- Could not open workbook: {exc}", ""])
            continue
        lines.append(f"- Sheets: {', '.join(wb_values.sheetnames)}")
        for sheet_name in wb_values.sheetnames:
            ws_values = wb_values[sheet_name]
            ws_formulas = wb_formulas[sheet_name]
            lines.extend(
                [
                    "",
                    f"### Sheet: {sheet_name}",
                    f"- Dimensions: rows={ws_values.max_row}, cols={ws_values.max_column}",
                ]
            )
            preview_rows = []
            max_rows = min(ws_values.max_row, 12)
            max_cols = min(ws_values.max_column, 12)
            for row in ws_values.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=True):
                preview_rows.append([safe_cell_value(value, 80) for value in row])
            if preview_rows:
                lines.append("- Top-left preview:")
                for idx, row in enumerate(preview_rows, start=1):
                    lines.append(f"  - R{idx}: {row}")

            formula_cells = []
            for row in ws_formulas.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_cells.append(f"{cell.coordinate}={safe_cell_value(value, 140)}")
                    if len(formula_cells) >= 25:
                        break
                if len(formula_cells) >= 25:
                    break
            if formula_cells:
                lines.append("- Formula examples:")
                for formula in formula_cells:
                    lines.append(f"  - {formula}")
        lines.append("")

    output = task_dir / "workbook_summary.md"
    output.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output.name


def generate_csv_profile(task_dir: Path) -> str | None:
    csv_paths = sorted(path for path in task_dir.glob("*.csv") if path.name != "csv_profile.md")
    if not csv_paths:
        return None
    lines = [
        "# CSV Profile",
        "",
        "This file was generated by the DSBench harness to expose CSV structure before agent execution.",
        "",
    ]
    for csv_path in csv_paths:
        lines.extend([f"## {csv_path.name}", ""])
        try:
            with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                rows = [row for _, row in zip(range(12), reader)]
        except UnicodeDecodeError:
            with csv_path.open("r", encoding="latin-1", newline="") as f:
                reader = csv.reader(f)
                rows = [row for _, row in zip(range(12), reader)]
        except Exception as exc:
            lines.extend([f"- Could not read CSV: {exc}", ""])
            continue
        if not rows:
            lines.extend(["- Empty CSV", ""])
            continue
        lines.append(f"- Header/first row: {[safe_cell_value(value, 80) for value in rows[0]]}")
        lines.append("- First rows:")
        for idx, row in enumerate(rows[1:8], start=2):
            lines.append(f"  - R{idx}: {[safe_cell_value(value, 80) for value in row[:12]]}")
        lines.append("")
    output = task_dir / "csv_profile.md"
    output.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output.name


def parse_question_options(question: str) -> dict[str, str]:
    options: dict[str, str] = {}
    patterns = [
        r"(?m)^\s*([A-I])[\).:]\s+(.+?)(?=^\s*[A-I][\).:]|\Z)",
        r"(?m)^\s*([A-I])\s+-\s+(.+?)(?=^\s*[A-I]\s+-|\Z)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, question, flags=re.S):
            key = match.group(1).strip().upper()
            value = " ".join(match.group(2).split())
            if value:
                options[key] = value
        if options:
            break
    return options


def generate_question_schema(task_dir: Path, question_name: str, question: str, answer) -> str:
    answer_format = "multiple_choice_letter" if is_multiple_choice_answer(answer) else "open_answer"
    schema = {
        "question_file": f"{question_name}.txt",
        "answer_format": answer_format,
        "requires_option_letter": answer_format == "multiple_choice_letter",
        "options": parse_question_options(question),
        "output_file": "final_answer.txt",
        "output_rules": [
            "Write only the final answer string.",
            "Do not include explanation, markdown, code, or debug output.",
        ],
    }
    if answer_format == "multiple_choice_letter":
        schema["output_rules"].append("Output exactly one uppercase option letter.")
    output = task_dir / "question_schema.json"
    output.write_text(json.dumps(schema, ensure_ascii=False, indent=2), encoding="utf-8")
    return output.name


def build_prompt(
    sample: dict,
    question_name: str,
    introduction: str,
    question: str,
    support_files: list[str],
    answer,
) -> str:
    support_block = "\n".join(f"- {name}" for name in support_files) if support_files else "- (none)"
    extra_instructions = ""
    if is_multiple_choice_answer(answer):
        extra_instructions = (
            "\n6. This task is a multiple-choice question.\n"
            "7. First compute the underlying result internally, then map that result to exactly one option letter from the question.\n"
            "8. Do not write any intermediate numeric value, percentage, currency amount, team name, free-text answer, or explanation to `final_answer.txt`.\n"
            "9. For this task, `final_answer.txt` must contain exactly one uppercase option letter such as `A`, `B`, `C`, `D`, `E`, `F`, `G`, `H`, or `I`.\n"
            "10. If you compute a value such as `728050`, `$34,274,780`, or `123.4%`, do not write that value to `final_answer.txt`; write only its corresponding option letter.\n"
            "11. If `final_answer.txt` contains anything other than one uppercase option letter, the answer is invalid."
        )
    return PROMPT_TEMPLATE.format(
        challenge_id=sample["id"],
        challenge_name=sample["name"],
        year=sample.get("year", ""),
        question_name=question_name,
        introduction=introduction.strip(),
        question=question.strip(),
        support_files=support_block,
        extra_instructions=extra_instructions,
    )


def main():
    args = parse_args()
    dsbench_root = Path(args.dsbench_root)
    source_zip = Path(args.source_zip) if args.source_zip else dsbench_root / "data_analysis" / "data_old.zip"
    meta_file = Path(args.meta_file) if args.meta_file else dsbench_root / "data_analysis" / "data.json"
    output_data_dir = Path(args.output_data_dir)
    taskset_file = Path(args.taskset_file)

    if not source_zip.exists():
        raise FileNotFoundError(f"DSBench source zip not found: {source_zip}")
    if not meta_file.exists():
        raise FileNotFoundError(f"DSBench metadata file not found: {meta_file}")

    samples = {
        sample["id"]: sample
        for sample in (json.loads(line) for line in meta_file.read_text(encoding="utf-8").splitlines() if line.strip())
    }
    preset = CURATED_PRESETS[args.preset]

    created_task_ids = []
    selection_notes = []

    with ZipFile(source_zip) as zf:
        names = set(zf.namelist())
        for block in preset["selections"]:
            sample = samples[block["challenge_id"]]
            question_to_answer = dict(zip(sample["questions"], sample["answers"]))
            challenge_dir = f"data/{sample['id']}/"
            intro_name = f"{challenge_dir}introduction.txt"
            if intro_name not in names:
                raise FileNotFoundError(f"Missing {intro_name} in {source_zip}")
            introduction = zf.read(intro_name).decode("utf-8", errors="ignore")

            support_members = []
            for member in sorted(names):
                if not member.startswith(challenge_dir):
                    continue
                basename = member.split("/")[-1]
                if not basename or basename.startswith("._") or basename == ".DS_Store":
                    continue
                if basename == "introduction.txt":
                    continue
                if basename.startswith("question") and basename.endswith(".txt"):
                    continue
                support_members.append(member)

            for question_name in block["question_names"]:
                question_file_name = f"{challenge_dir}{question_name}.txt"
                if question_file_name not in names:
                    raise FileNotFoundError(f"Missing {question_file_name} in {source_zip}")
                question = zf.read(question_file_name).decode("utf-8", errors="ignore")
                answer = question_to_answer[question_name]

                task_name = f"{args.task_prefix}_{sample['id']}_{sanitize_name(question_name)}"
                task_dir = output_data_dir / task_name
                if task_dir.exists():
                    if not args.overwrite:
                        raise FileExistsError(f"{task_dir} already exists; rerun with --overwrite to replace it")
                    shutil.rmtree(task_dir)
                task_dir.mkdir(parents=True, exist_ok=True)

                for member in support_members:
                    target = task_dir / Path(member).name
                    with zf.open(member) as src, target.open("wb") as dst:
                        shutil.copyfileobj(src, dst)

                (task_dir / "introduction.txt").write_text(introduction, encoding="utf-8")
                (task_dir / f"{question_name}.txt").write_text(question, encoding="utf-8")

                support_file_names = [Path(member).name for member in support_members]
                support_file_names.extend(["introduction.txt", f"{question_name}.txt"])
                generated_support_files = [
                    item
                    for item in [
                        generate_workbook_summary(task_dir),
                        generate_csv_profile(task_dir),
                        generate_question_schema(task_dir, question_name, question, answer),
                    ]
                    if item
                ]
                support_file_names.extend(generated_support_files)

                prompt_payload = {
                    "prompt": build_prompt(
                        sample,
                        question_name,
                        introduction,
                        question,
                        sorted(support_file_names),
                        answer,
                    ),
                    "data_source_type": "2=open source data",
                }
                metadata_payload = {
                    "benchmark": "DSBench",
                    "benchmark_track": "data_analysis",
                    "challenge_id": sample["id"],
                    "challenge_name": sample["name"],
                    "question_name": question_name,
                    "question": question,
                    "introduction": introduction,
                    "answer": answer_to_text(answer),
                    "answer_raw": answer,
                    "answer_format": "multiple_choice_letter" if is_multiple_choice_answer(answer) else "open_answer",
                    "year": sample.get("year", ""),
                    "source_zip": str(source_zip),
                    "selection_preset": args.preset,
                    "selection_reason": block["reason"],
                }

                (task_dir / "prompt.json").write_text(json.dumps(prompt_payload, ensure_ascii=False), encoding="utf-8")
                (task_dir / "metadata.json").write_text(
                    json.dumps(metadata_payload, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                created_task_ids.append(task_name)

            selection_notes.append(
                {
                    "challenge_id": sample["id"],
                    "challenge_name": sample["name"],
                    "reason": block["reason"],
                    "question_names": block["question_names"],
                }
            )

    taskset_file.parent.mkdir(parents=True, exist_ok=True)
    manifest = {
        f"{args.task_prefix}_{args.preset}": created_task_ids,
        f"{args.task_prefix}_{args.preset}_meta": {
            "description": preset["description"],
            "size": len(created_task_ids),
            "selection_notes": selection_notes,
        },
    }
    taskset_file.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Prepared {len(created_task_ids)} DSBench data-analysis tasks for preset {args.preset}")
    print(f"Task set manifest: {taskset_file}")
    print(f"Task ids: {created_task_ids}")


if __name__ == "__main__":
    main()
